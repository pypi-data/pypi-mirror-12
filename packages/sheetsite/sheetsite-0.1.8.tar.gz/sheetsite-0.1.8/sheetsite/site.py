from collections import OrderedDict
import json
import os
import re
import sys
from sheetsite.names import normalize_name

class Site(object):

    def __init__(self, spreadsheet, geocache_filename, censor=True):
        self.workbook = spreadsheet
        self.geocache_filename = geocache_filename
        self.censor = censor

    def save_local(self, output_file):
        ext = '-'
        if output_file is not None:
            _, ext = os.path.splitext(output_file)
            ext = ext.lower()

        if ext == ".xls" or ext == ".xlsx":
            return self.save_to_excel(output_file)
        elif ext == ".json" or ext == '-':
            return self.save_to_json(output_file)

        print("Unknown extension", ext)
        return False

    def process_cells(self, rows):
        rows = self.clean_cells(rows)
        rows = self.add_location(rows)
        return rows
        

    def save_to_excel(self, output_file):
        from openpyxl import Workbook
        wb = Workbook()
        first = True
        for sheet in self.workbook.worksheets():
            if first:
                ws = wb.active
                first = False
            else:
                ws = wb.create_sheet()
            ws.title = sheet.title
            rows = self.process_cells(sheet.get_all_values())
            for r, row in enumerate(rows):
                for c, cell in enumerate(row):
                    ws.cell(row=r+1, column=c+1).value = cell
        wb.save(output_file)
        return True

    def save_to_json(self, output_file):
        result = OrderedDict()
        order = result['names'] = []
        sheets = result['tables'] = OrderedDict()
        for sheet in self.workbook.worksheets():
            order.append(sheet.title)
            ws = sheets[sheet.title] = OrderedDict()
            vals = self.process_cells(sheet.get_all_values())
            if len(vals)>0:
                columns = vals[0]
                rows = vals[1:]
                ws['columns'] = columns
                ws['rows'] = [OrderedDict(zip(columns, row)) for row in rows]
            else:
                ws['columns'] = []
                ws['rows'] = []
        if output_file == None:
            print json.dumps(result, indent=2)
        else:
            with open(output_file, 'w') as f:
                json.dump(result, f, indent=2)
        return True

    def clean_cells(self, vals):
        if len(vals) == 0:
            return vals

        hide_column = {}
        for idx, cell in enumerate(vals[0]):
            if cell is None or len(cell)==0 or cell[0] == '(':
                hide_column[idx] = True

        results = []

        for ridx, row in enumerate(vals):
            result = []
            for idx, cell in enumerate(row):
                if idx in hide_column:
                    continue
                if cell is not None:
                    cell = re.sub(r'\(\(.*\)\)','', cell)
                    cell = re.sub(r'[\n\r]+$','', cell)
                    cell = re.sub(r'^[\t \n\r]+$','', cell)
                result.append(cell)
            results.append(result)

        return results

    def add_location(self, vals):
        if len(vals) == 0:
            return vals

        have_address = False
        have_fill_in = False
        fill_in = []
        key_id = 0
        for idx, cell in enumerate(vals[0]):
            if normalize_name(cell) == 'address':
                key_id = idx
                have_address = True
            if cell is not None and len(cell)>0 and cell[0] == '[':
                have_fill_in = True
                vals[0][idx] = cell[1:-1]
                fill_in.append([normalize_name(vals[0][idx]), idx])
        if not(have_fill_in) or not(have_address):
            return vals
        from sheetsite.geocache import GeoCache
        cache = GeoCache(self.geocache_filename)
        cache.find_all(vals[1:], key_id, fill_in)
        return vals
